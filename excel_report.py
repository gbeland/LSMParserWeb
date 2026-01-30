import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
import re
from config import *

def sanitize_value(val):
    """Removes illegal characters that openpyxl cannot handle."""
    if isinstance(val, str):
        # XML 1.0 valid chars: #x9 | #xA | #xD | [#x20-#xD7FF] | ...
        # But openpyxl is stricter. Let's just strip vertical tabs (0x0B), form feeds (0x0C), etc.
        # Common issue is control characters.
        # This regex removes control characters except newlines/tabs
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)
    return val

def set_cell(sheet, row, col, value):
    """Helper to set cell value safely."""
    try:
        sheet.cell(row=row, column=col, value=sanitize_value(value))
    except IllegalCharacterError:
        # Fallback to string repr or just empty if really bad
        try:
            sheet.cell(row=row, column=col, value=str(value).encode('ascii', 'ignore').decode())
        except:
            sheet.cell(row=row, column=col, value="[Invalid Char]")
    return sheet.cell(row=row, column=col)

def style_cell(cell, fill_color=None, bold=False, align=None):
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if bold:
        cell.font = Font(bold=True)
    if align:
        cell.alignment = Alignment(horizontal=align)

def auto_adjust_columns(sheet):
    """Simple auto-width adjustment."""
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

def create_excel_report(wb, sbb_data, cab_data, layout_data, cab_stats, source_filename=""):
    """Generates the SBBInfo, CabInfo, and CabLayouts sheets using processed data."""
    
    # Refresh Sheets
    if "SBBInfo" in wb.sheetnames: del wb["SBBInfo"]
    ws_sbb = wb.create_sheet("SBBInfo")
    
    if "CabInfo" in wb.sheetnames: del wb["CabInfo"]
    ws_cab = wb.create_sheet("CabInfo")

    if "CabLayouts" in wb.sheetnames: del wb["CabLayouts"]
    ws_layout = wb.create_sheet("CabLayouts")
    
    # Write Source Header
    if source_filename:
        for ws in [ws_sbb, ws_cab, ws_layout]:
             ws.merge_cells('A1:E1')
             cell = ws.cell(row=1, column=1, value=f"Source File: {source_filename}")
             style_cell(cell, bold=True)
             # Offset other rows by 1?
             # Actually, simpler to just start headers at row 2 now.
             
    start_row = 2 if source_filename else 1

    # --- SBB Info ---
    headers_sbb = [
        "Setting", "SBox IP", "Model", "Serial Number", "Main Version", 
        "Additional Versions", "Device Name (0x67)", "IMEI",
        "Input Source", "Power Status", "Status Error",
        "MAC Address", "Network IP Mode",
        "Group 1 IP", "Group 2 IP", 
        "Group 3 IP", "Group 4 IP", "Video Wall Mode", "Video Offset", "Cabinet Layout",
        "SBox Output Resolution", "Cabinet Resolution"
    ]
    for r, h in enumerate(headers_sbb):
        cell = ws_sbb.cell(row=start_row + r, column=1, value=h)
        style_cell(cell, bold=True)
        
    col_idx = 2
    # row_data_start is not strictly needed if we just offset
    for sbb in sbb_data:
        sbb['excel_col'] = col_idx # Store for layout updating
        
    for sbb in sbb_data:
        sbb['excel_col'] = col_idx # Store for layout updating
        
        # Helper for transposed rows
        def set_sbb(r_offset, val):
             return set_cell(ws_sbb, start_row + r_offset - 1, col_idx, val)

        c = set_sbb(1, sbb['name_header'])
        style_cell(c, bold=True)
        
        set_sbb(2, sbb['ip'])
        set_sbb(3, sbb['model'])
        set_sbb(4, sbb['sn'])
        set_sbb(5, sbb['fw_main'])
        c = set_sbb(6, sbb['fw_add'])
        c.alignment = Alignment(wrap_text=True, vertical='top')
        set_sbb(7, sbb['sbb_name'])
        set_sbb(8, sbb.get('imei', ''))
        
        # New Fields
        set_sbb(9, sbb.get('input', ''))
        set_sbb(10, sbb.get('power', ''))
        set_sbb(11, sbb.get('status', ''))
        set_sbb(12, sbb.get('mac', ''))
        set_sbb(13, sbb.get('ip_mode', ''))
        
        # Groups (Rows 14-17)
        for g_num, ip_str in sbb['groups'].items():
            set_sbb(13+g_num, ip_str)
            
        c = set_sbb(18, sbb['vw_mode'])
        if sbb['vw_mode'] == "On": style_cell(c, fill_color=COLOR_GREEN)
        
        set_sbb(19, sbb.get('video_offset', ''))
        set_sbb(20, sbb.get('layout_str', ''))
        set_sbb(21, sbb.get('res_sbb', ''))
        set_sbb(22, sbb.get('res_cab', ''))
        
        col_idx += 1

    # --- Cab Info ---
    # Summary Section
    # Display summary at start_row
    
    # Headers for Summary
    set_cell(ws_cab, start_row, 1, "Cabinet Summary")
    style_cell(ws_cab.cell(row=start_row, column=1), bold=True)
    
    summary_headers = ["Total Cabs", "Most Common Model", "Most Common Main FW", "Most Common FPGA FW", "Avg Temp", "Max Temp", "Min Temp"]
    for i, h in enumerate(summary_headers):
        c = set_cell(ws_cab, start_row+1, i+1, h)
        style_cell(c, bold=True)
        
    set_cell(ws_cab, start_row+2, 1, cab_stats.get('count', 0))
    set_cell(ws_cab, start_row+2, 2, cab_stats.get('mode_model', ''))
    set_cell(ws_cab, start_row+2, 3, cab_stats.get('mode_fw_main', ''))
    set_cell(ws_cab, start_row+2, 4, cab_stats.get('mode_fw_fpga', ''))
    set_cell(ws_cab, start_row+2, 5, cab_stats.get('temp_avg', ''))
    c = set_cell(ws_cab, start_row+2, 6, cab_stats.get('temp_max', ''))
    if cab_stats.get('temp_max') is not None: style_cell(c, fill_color=COLOR_PINK)
    set_cell(ws_cab, start_row+2, 7, cab_stats.get('temp_min', ''))

    auto_adjust_columns(ws_cab)
    
    # Group by SBox IP
    # Get unique IPs preserving order
    unique_ips = []
    seen = set()
    for c in cab_data:
        if c['sbb_ip'] not in seen:
            unique_ips.append(c['sbb_ip'])
            seen.add(c['sbb_ip'])
            
    list_start_row = start_row + 5
    row_cab = list_start_row
    
    headers_cab = [
        "SBox IP", "Group IP", "Cab ID", "Serial#", "Model#", 
        "Cabinet FW", "Cabinet FPGA", "Temp(c)", "BackLt", 
        "Cab RGB CC", "Mod RGB CC", "Pix RGB CC", "Seam Cor",
        "Video Loc"
    ]
    
    for s_ip in unique_ips:
        # Filter Data
        sbox_cabs = [c for c in cab_data if c['sbb_ip'] == s_ip]
        cab_count = len(sbox_cabs)
        
        # Section Header
        set_cell(ws_cab, row_cab, 1, f"SBox: {s_ip} (Total Cabinets: {cab_count})")
        style_cell(ws_cab.cell(row=row_cab, column=1), bold=True, fill_color="CCCCCC") # Gray background
        row_cab += 1
        
        # Table Headers
        for c, h in enumerate(headers_cab, 1):
            ws_cab.cell(row=row_cab, column=c, value=h)
            style_cell(ws_cab.cell(row=row_cab, column=c), bold=True)
        row_cab += 1
        
        for cab in sbox_cabs:
            set_cell(ws_cab, row_cab, 1, cab['sbb_ip'])
            set_cell(ws_cab, row_cab, 2, cab['group_ip'])
            set_cell(ws_cab, row_cab, 3, cab['cid'])
            set_cell(ws_cab, row_cab, 4, cab['sn'])
            
            c = set_cell(ws_cab, row_cab, 5, cab['model'])
            if cab['model'] != cab_stats['mode_model']: style_cell(c, fill_color=COLOR_PINK)
            
            c = set_cell(ws_cab, row_cab, 6, cab['fw_main'])
            if cab['fw_main'] != cab_stats['mode_fw_main']: style_cell(c, fill_color=COLOR_PINK)
            
            c = set_cell(ws_cab, row_cab, 7, cab['fw_fpga'])
            if cab['fw_fpga'] != cab_stats['mode_fw_fpga']: style_cell(c, fill_color=COLOR_PINK)
            
            if cab['temp'] is not None:
                 c = set_cell(ws_cab, row_cab, 8, cab['temp'])
                 if cab_stats['temp_max'] is not None and cab['temp'] == cab_stats['temp_max']:
                     style_cell(c, fill_color=COLOR_PINK)
                 elif cab['temp'] > 59: 
                     style_cell(c, fill_color=COLOR_YELLOW)
            
            set_cell(ws_cab, row_cab, 9, cab['backlight'])
            
            # Binary Settings
            _set_binary_cell(ws_cab, row_cab, 10, cab['cc_cab'])
            _set_binary_cell(ws_cab, row_cab, 11, cab['cc_mod'])
            _set_binary_cell(ws_cab, row_cab, 12, cab['cc_pix'])
            _set_binary_cell(ws_cab, row_cab, 13, cab['seam'])
            set_cell(ws_cab, row_cab, 14, cab['video_location'])
            
            row_cab += 1
            
        row_cab += 1 # Spacer row

    # --- Layouts ---
    _generate_layout_sheet(ws_layout, ws_sbb, layout_data, sbb_data)
    
    auto_adjust_columns(ws_sbb)
    auto_adjust_columns(ws_cab)

def _set_binary_cell(ws, row, col, val):
    c = set_cell(ws, row, col, val)
    if val == "On": style_cell(c, fill_color=COLOR_GREEN)

def _generate_layout_sheet(ws_layout, ws_sbb, layout_data, sbb_data):
    row_layout = 1
    col_layout_start = 1
    
    # Needs to match sbb_data order? or sorted IPs?
    # layout_data is dict sbb_ip -> cabs
    
    # Map IP to SBB Data entry for column update
    sbb_map = {s['ip']: s for s in sbb_data}
    
    for s_ip in sorted(layout_data.keys()):
        cabs = layout_data[s_ip]
        if not cabs: continue
        
        xs = sorted(list(set(c['x_sbb'] for c in cabs)))
        ys = sorted(list(set(c['y_sbb'] for c in cabs)))
        
        # Header
        ws_layout.cell(row=row_layout, column=col_layout_start, value=f"SBox: {s_ip}")
        style_cell(ws_layout.cell(row=row_layout, column=col_layout_start), bold=True)
        row_layout += 1
        
        start_row = row_layout
        for cab in cabs:
            try:
                col_idx = xs.index(cab['x_sbb'])
                row_idx = ys.index(cab['y_sbb'])
                
                
                val = f"Grp: {cab['group']}\nID: {cab['cid']}\n{cab['x_sbb']}x{cab['y_sbb']}"
                c = set_cell(ws_layout, start_row+row_idx, col_layout_start+col_idx, val)
                c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                
                color = None
                if cab['group'] == 1: color = COLOR_GREEN
                elif cab['group'] == 2: color = COLOR_VIOLET
                elif cab['group'] == 3: color = COLOR_BLUE
                elif cab['group'] == 4: color = COLOR_TAN
                if color: style_cell(c, fill_color=color)
            except: pass

        # Update SBB Sheet Summary using stored col index
        if s_ip in sbb_map and 'excel_col' in sbb_map[s_ip]:
            found_col = sbb_map[s_ip]['excel_col']
            cw, ch = cabs[0]['w'], cabs[0]['h']
            cols, rows = len(xs), len(ys)
            
            set_cell(ws_sbb, 20, found_col, f"{cols} x {rows}")
            set_cell(ws_sbb, 21, found_col, f"{cols*cw} x {rows*ch}")
            set_cell(ws_sbb, 22, found_col, f"{cw} x {ch}")
        
        row_layout += len(ys) + 2
