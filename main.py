import sys
import os
import tkinter as tk
from tkinter import filedialog
import openpyxl
import zipfile
import traceback
import xml.etree.ElementTree as ET
import re
import glob
import logging

# Configure Logger
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("LSMParser.log", mode='w'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("LSMParser")
# Mute other loggers to avoid noise if necessary, or set their level
logging.getLogger("openpyxl").setLevel(logging.WARNING)

from config import VERSION
from mdc_parser import hex_string_to_list
from data_processor import process_logs
from excel_report import create_excel_report
from html_report import create_html_report
from raw_parser import read_xlsx_raw
from com_parser import read_xlsx_via_com

def print_help():
    print("\n--- LSM Parser Help ---")
    print("Usage:")
    print("  1. Run without arguments to open file selection dialog.")
    print("  2. Drag and drop an .xlsx file onto the script.")
    print("  3. Run from command line: python main.py <path_to_file>")
    print("\nRequirements:")
    print("  - Input file must be a valid .xlsx file.")
    print("-----------------------------\n")

def print_report_summary(sbb_data, cab_data):
    print("\n" + "="*40)
    print("       PARSING SUMMARY")
    print("="*40)
    
    if not sbb_data:
        print("No SBoxes found.")
        return

    for i, sbb in enumerate(sbb_data, 1):
        print(f"\n[SBox {i}]")
        print(f"  IP Address    : {sbb.get('ip', 'N/A')}")
        print(f"  Model         : {sbb.get('model', 'Unknown')}")
        print(f"  Serial Number : {sbb.get('sn', 'Unknown')}")
        print(f"  MAC Address   : {sbb.get('mac', 'Unknown')}")
        
        # Count cabs for this sbb
        cabs_count = len([c for c in cab_data if c['sbb_ip'] == sbb['ip']])
        print(f"  Cabinets      : {cabs_count}")
        print(f"  Resolution    : {sbb.get('res_sbb', 'Unknown')}")
        
    print("\n" + "="*40 + "\n")



def select_inputs():
    """Opens a dialog to select files or a directory."""
    selection = []
    
    def select_files():
        files = filedialog.askopenfilenames(
            title="Select LSM Log Spreadsheets",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if files:
            selection.extend(files)
            root.quit()
            
    def select_folder():
        folder = filedialog.askdirectory(title="Select Log Directory")
        if folder:
            selection.append(folder)
            root.quit()
    
    root = tk.Tk()
    root.title("LSM Input")
    
    # Center window simple calculation (optional, or just let OS handle)
    # Keeping it simple for now
    
    lbl = tk.Label(root, text="Select Input Source:", padx=20, pady=10)
    lbl.pack()
    
    btn_files = tk.Button(root, text="Select File(s)", command=select_files, width=20)
    btn_files.pack(pady=5)
    
    btn_folder = tk.Button(root, text="Select Folder", command=select_folder, width=20)
    btn_folder.pack(pady=5)
    
    # Add a cancel/close listener
    root.protocol("WM_DELETE_WINDOW", root.quit)
    
    root.mainloop()
    try:
        root.destroy()
    except:
        pass
    return selection

def expand_paths(paths):
    """Expands directories in the path list to .xlsx files."""
    expanded = []
    for p in paths:
        if os.path.isdir(p):
            # Expand directory
            dir_contents = glob.glob(os.path.join(p, "*.xlsx"))
            if dir_contents:
                expanded.extend(dir_contents)
                print(f"Expanded directory '{p}' to {len(dir_contents)} files.")
        else:
            expanded.append(p)
    return expanded

def analyze_file_logic(file_path):
    """
    Core analysis logic that can be reused by GUI.
    Returns tuple (excel_path, html_path) or None on failure.
    """
    print("\n" + "-"*40)
    logger.info(f"Opening: {file_path}")
    print("-" * 40)
    
    if not file_path or not os.path.exists(file_path):
        logger.error(f"File does not exist: {file_path}")
        print("File does not exist.")
        return None

    # Check for NASCA DRM
    is_drm = False
    try:
        with open(file_path, 'rb') as f:
            header = f.read(50)
            if b"NASCA DRM FILE" in header:
                is_drm = True
                logger.warning(f"File '{file_path}' is encrypted with NASCA DRM.")
    except Exception as e:
        logger.debug(f"Pre-check file read failed: {e}")
        
    use_raw_data = False
    lsm_rows = []
    
    # Strategy 1: If DRM, try COM immediately
    if is_drm:
        print("Notice: DRM Encryption detected. Attempting to use Excel COM to decrypt...")
        com_data = read_xlsx_via_com(file_path)
        if com_data:
            print("Success: Decrypted via Excel COM.")
            lsm_rows = com_data
            use_raw_data = True
        else:
             logger.error("File is encrypted with NASCA DRM and Excel COM failed.")
             print(f"Error: The file '{file_path}' is encrypted with NASCA DRM.")
             print("Excel Automation failed or Excel is not installed.")
             print("Please decrypt the file using your corporate security tool (e.g. Fasoo/Safer) before processing.")
             return None

    # Strategy 2: Standard OpenPyXL
    if not use_raw_data:
        try:
            wb_in = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet1 = wb_in.worksheets[0]
        except Exception as e:
            # Check for known issues
            err_msg = str(e)
            is_bad_zip = isinstance(e, zipfile.BadZipFile) or "BadZipFile" in str(type(e))
            is_openpyxl_error = "Fill() takes no arguments" in err_msg or "expected <class 'openpyxl.styles.fills.Fill'>" in err_msg
            
            if is_bad_zip:
                 logger.error("BadZipFile: Invalid .xlsx format.")
            elif is_openpyxl_error:
                 logger.warning(f"OpenPyXL compatibility issue ({e}).")
            else:
                 logger.error(f"Error opening file with openpyxl: {e}", exc_info=True)
                 print(f"Error opening file with openpyxl: {e}")

            # Strategy 3: OpenPyXL failed, try COM Fallback
            print("Attempting fallback to Excel COM...")
            com_data = read_xlsx_via_com(file_path)
            if com_data:
                print("Success: Read via Excel COM.")
                lsm_rows = com_data
                use_raw_data = True
            else:
                # Strategy 4: COM failed, try Raw XML parsing
                logger.info("Falling back to raw XML parsing...")
                print("Notice: Standard parser failed, falling back to raw XML parsing...")
                raw_data = read_xlsx_raw(file_path)
                if not raw_data:
                    return None
                
                use_raw_data = True
                lsm_rows = raw_data


    # 2. Read Logging Data
    lsm_logs = []
    
    print("Importing Data...")
    
    iterator = lsm_rows if 'use_raw_data' in locals() and use_raw_data else sheet1.iter_rows(min_row=1, values_only=True)
    
    for row in iterator:

        if not row or len(row) < 5:
            continue
            
        mdc_resp = str(row[4]) if row[4] else ""
        if "AA FF" not in mdc_resp:
            continue
            
        dev_ip = str(row[3]) if row[3] else ""
        
        parts = hex_string_to_list(mdc_resp)
        if len(parts) < 3:
            continue
            
        try:
            dev_id = int(parts[2], 16)
            lsm_logs.append({
                'ip': dev_ip,
                'id': dev_id,
                'resp': mdc_resp,
                'parts': parts
            })
        except ValueError:
            continue
    
    print(f"Data import complete. Found {len(lsm_logs)} raw MDC frames.")
    logger.info(f"Raw logs found: {len(lsm_logs)}")
    
    # Identify SBoxes and Cabinets
    sbb_list = {}
    cab_list = {}
    
    for entry in lsm_logs:
        key = (entry['ip'], entry['id'])
        if entry['id'] == 1:
            if key not in sbb_list: sbb_list[key] = []
            sbb_list[key].append(entry)
        else:
            if key not in cab_list: cab_list[key] = []
            cab_list[key].append(entry)

    print(f"SBoxes found: {len(sbb_list)}")
    print(f"Cabinets found: {len(cab_list)}")
    logger.debug(f"SBoxes: {len(sbb_list)}, Cabinets: {len(cab_list)}")
    
    # Close input file if it was opened
    if 'wb_in' in locals():
        wb_in.close()

    # 3. Process Data
    print("Processing Data...")
    logger.info("Processing log data...")
    sbb_data, cab_data, layout_data, cab_stats = process_logs(sbb_list, cab_list)

    # 4. Generate Reports
    base_name = os.path.splitext(file_path)[0]
    source_filename = os.path.basename(file_path)
    
    # Excel
    print("Generating Excel Report...")
    wb_out = openpyxl.Workbook()
    
    # Remove default sheet if we are going to create our own, or use it
    if "Sheet" in wb_out.sheetnames:
        del wb_out["Sheet"]
        
    create_excel_report(wb_out, sbb_data, cab_data, layout_data, cab_stats, source_filename=source_filename)
    xlsx_path = f"{base_name}-Parsed.xlsx"
    wb_out.save(xlsx_path)
    print(f"Saved: {xlsx_path}")
    logger.info(f"Saved Excel: {xlsx_path}")
    
    # HTML
    print("Generating HTML Report...")
    html_path = f"{base_name}-Parsed.html"
    create_html_report(sbb_data, cab_data, layout_data, cab_stats, html_path, source_filename=source_filename)
    print(f"Saved: {html_path}")
    logger.info(f"Saved HTML: {html_path}")
    
    # Summary
    print_report_summary(sbb_data, cab_data)
    
    print("Done.")

    print("Analysis finished for file.")
    logger.info("Analysis finished.")
    
    return xlsx_path, html_path

def analyze_file(file_path):
    """CLI wrapper for analysis logic."""
    try:
        analyze_file_logic(file_path)
    except Exception as e:
        # Catch exceptions here for CLI robustness, though they might be caught in main loop too
        raise e

def main():
    print(f"LSM Parser v{VERSION}")
    logger.info(f"Starting LSM Parser v{VERSION}")
    
    # 1. Gather files
    raw_inputs = []
    
    if len(sys.argv) > 1:
        # Check for standard flags early
        first_arg = sys.argv[1].lower()
        if first_arg in ["--help", "-h", "/?"]:
            print_help()
            return
        if first_arg in ["--version", "-v"]:
            print(f"LSM Parser v{VERSION}")
            return

        # Check arguments for wildcards and dirs
        for arg in sys.argv[1:]:
            # Use glob to pattern match
            matched = glob.glob(arg)
            if matched:
                raw_inputs.extend(matched)
            else:
                # If glob doesn't find anything, try adding raw arg
                print(f"Warning: No files found matching '{arg}'")
    
    else:
        # No args, use GUI picker
        selected = select_inputs()
        if selected:
            raw_inputs.extend(selected)
        else:
            print("No inputs selected.")
            print_help()
            
            # Wait for exit if double clicked
            print("\nPress Return to exit...")
            input()
            return

    # Expand directories and filter
    files_to_process = expand_paths(raw_inputs)

    # Remove duplicates and filter ignored files
    unique_files = sorted(list(set(files_to_process)))
    files_to_process = []
    for f in unique_files:
        if not os.path.isfile(f):
            # Skip directories quietly
            continue
            
        if not f.lower().endswith(".xlsx"):
            print(f"Skipping non-xlsx file: {f}")
            continue

        if f.lower().endswith("-parsed.xlsx"):
            print(f"Skipping already parsed file: {f}")
            logger.info(f"Skipping already parsed file: {f}")
            continue
        files_to_process.append(f)
    
    if not files_to_process:
        # Smart Stitching: Try to see if arguments form a valid filename (unquoted path support)
        if len(sys.argv) > 1:
            reconstructed_path = " ".join(sys.argv[1:])
            # We also check if it might be a glob pattern that just didn't match (less likely if os expanded it, 
            # but if they passed unquoted wildcard that expanded to nothing? No, shell handles that.
            # Just check if it's a file.)
            if os.path.exists(reconstructed_path) and os.path.isfile(reconstructed_path):
                 print(f"Notice: Detected unquoted path '{reconstructed_path}'. Processing...")
                 logger.info(f"Smart stitch detected file: {reconstructed_path}")
                 files_to_process.append(reconstructed_path)
    
    if not files_to_process:
        print("No files to process.")
    else:
        print(f"Files queued for processing: {len(files_to_process)}")
        for f in files_to_process:
            try:
                analyze_file(f)
            except Exception as e:
                logger.critical(f"CRITICAL ERROR processing {f}: {e}", exc_info=True)
                print(f"CRITICAL ERROR processing {f}: {e}")
                traceback.print_exc()



if __name__ == "__main__":
    main()
